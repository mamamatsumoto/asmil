import openpyxl


def is_empty(cell):
    return cell.value is None or not str(cell.value).strip()

def reading(uploaded_file):

    ### 読み込み
    ## excel data 取得
    wb = openpyxl.load_workbook(uploaded_file[0])
    ws = wb["template1"]

    # Header情報取得
    quote_header = []

    for row in ws.iter_rows(min_col=3, max_col=78, max_row=29):
        for cell in row:
            cell_value = cell.value
            if cell_value is not None:
                quote_header.append(cell_value)

    # 見積金額詳細情報取得
    quote_details = [] #見積表全体取得
    quote_start = 32 #見積変動開始行
    for i in ws.iter_rows(min_col=2, min_row=quote_start, max_col=78):
        if all(is_empty(c) for c in i):
            break
        values_original = [cell.value for cell in i]
        values_original_filtered = [j for j in values_original if j is not None]
        quote_details.append(values_original_filtered)

    quote_numbers = len(quote_details) #見積行数取得

    dt_row_p1 = 29 #見積項目数上限Page.1 // 29項目
    dt_row_p2 = 84 #見積項目数上限Page.2 // 84項目

    dt_col_list = []
    for k in quote_details:
        dt_col_list.append({
            "works":k[0],
            "volume" : '{:,.3f}'.format(k[2]),
            "unit_v" : k[3],
            "currency" : k[4],
            "unit_p" : '{:,.2f}'.format(k[5]),
            "amount" : '{:,.2f}'.format(k[6]),
            "ex_rate" : k[7],
            "amount_j" : '{:,}'.format(k[8]),
            "tax" : k[9],
        })
    dt_list_1 = dt_col_list[0:dt_row_p1]
    dt_list_2 = dt_col_list[dt_row_p1:dt_row_p2]
    dt_list_3 = dt_col_list[dt_row_p2:quote_numbers]


    # 金額計算
    tax_total_cal = []
    duty_free_total_cal = []
    for m in quote_details:
        if m[9] == '〇':
            tax_total_cal.append(m[8])
        else:
            duty_free_total_cal.append(m[8])

    grand_total = []

    duty_free_total = sum(duty_free_total_cal) # 免税品合計
    grand_total.append({"duty_free_total" : '{:,}'.format(duty_free_total)})

    tax_total = sum(tax_total_cal) # 課税品合計
    grand_total.append({"tax_total" : '{:,}'.format(tax_total)})

    gtotal = tax_total + duty_free_total #総合計
    grand_total.append({"gtotal" : '{:,}'.format(gtotal)})

    grandtotal_1 = []
    grandtotal_2 = []
    grandtotal_3 = []

    if quote_numbers <= dt_row_p1 - 4 or quote_numbers == dt_list_1:
        grandtotal_1 = grand_total
    elif quote_numbers <= dt_row_p2 - 4:
        grandtotal_2 = grand_total
    else:
        grandtotal_3 = grand_total


    # 付帯事項

    #付帯事項開始行
    remarks_start = quote_start + quote_numbers + 6 #備考欄開始地点
    remarks = []
    for row in ws.iter_rows(min_col=4, min_row=remarks_start):
        for cell in row:
            if all(is_empty(c) for c in row):
                break
            cell_value = cell.value
            if cell_value is not None:
                remarks.append(cell_value)

    remark1 = []
    remark2 = []
    remark3 = []

    if len(grandtotal_1) > 0:
        remark_space1 = (124.35 - (len(dt_list_1) * 4.95 + 23.43))/3.62
        remark1 = remarks[0:int(remark_space1)]
        remark2 = remarks[int(remark_space1):len(remarks)]
    elif len(grandtotal_2) > 0:
        remark_space2 = (272.25 - (len(dt_list_2) * 4.95 + 23.43))/3.62
        remark2 = remarks[0:int(remark_space2)]
        remark3 = remarks[int(remark_space2):len(remarks)]
    else:
        remark3 = remarks

    # Page設定
    page_2 = []
    page_3 = []
    page_2_dis = len(dt_list_2) + len(grandtotal_2) + len(remark2)
    page_3_dis = len(dt_list_3) + len(grandtotal_3) + len(remark3)
    if page_2_dis > 0:
        page_2.append(page_2_dis)
    if page_3_dis > 0:
        page_3.append(page_3_dis)
    wb.close()

    extracted_data=[
        quote_header,
        tax_total,
        duty_free_total,
        gtotal,
        grandtotal_1,
        grandtotal_2,
        grandtotal_3,
        dt_list_1,
        dt_list_2,
        dt_list_3,
        page_2,
        page_3,
        remark1,
        remark2,
        remark3,
    ]

    return extracted_data

def setting_context(extracted_data):

    quote_header = extracted_data[0]

    context = {
        'address' : quote_header[4:6],
        'issued_date': quote_header[1],
        'quote_no': quote_header[3],
        'section_1': quote_header[6],
        'section_2': quote_header[8],
        'pic': quote_header[10],
        'tel_no': quote_header[12],
        'fax_no': quote_header[14],
        'class1' : quote_header[19],
        'class2' : quote_header[21],
        'project_name' : quote_header[23],
        'item' : quote_header[25],
        'from' : quote_header[27],
        'to' : quote_header[30],
        'POL' : quote_header[33],
        'POD' : quote_header[36],
        'execute_from' : quote_header[39],
        'execute_upto' : quote_header[41],
        'validity' : quote_header[43],
        'packages' : '{:,}'.format(quote_header[46]),
        'max_l' : '{:,}'.format(quote_header[49]),
        't_gw' : '{:,}'.format(quote_header[52]),
        'max_w' : '{:,}'.format(quote_header[55]),
        't_mm' : '{:,.3f}'.format(quote_header[58]),
        'max_h' : '{:,}'.format(quote_header[61]),
        't_ft' : '{:,.3f}'.format(quote_header[64]),
        'max_wt' : '{:,}'.format(quote_header[67]),
        't_rt' : '{:,.3f}'.format(quote_header[70]),
        'cwt' : '{:,.1f}'.format(quote_header[73]),
        'scope' : quote_header[75:78],
        'tax_total' : '{:,}'.format(extracted_data[1]),
        'duty_free_total' : '{:,}'.format(extracted_data[2]),
        'page_2' : extracted_data[10],
        'page_3' : extracted_data[11],
        'dt_list_1' : extracted_data[7],
        'dt_list_2' : extracted_data[8],
        'dt_list_3' : extracted_data[9],
        'grand_total_1' : extracted_data[4],
        'grand_total_2' : extracted_data[5],
        'grand_total_3' : extracted_data[6],
        'gtotal' : '{:,}'.format(extracted_data[3]),
        'remark_1' : extracted_data[12],
        'remark_2' : extracted_data[13],
        'remark_3' : extracted_data[14],
        }
    return context