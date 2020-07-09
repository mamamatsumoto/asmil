import openpyxl
import pprint
import numpy as np
import os.path
import glob


temp_dir = "media/excel/Qe6WZMYcW9"
path = temp_dir + "/*"
file_list = glob.glob(path)

print(file_list[0])
print(os.path.join(path, "*.xlsx"))



wb = openpyxl.load_workbook("quotes/E2004076A01見積書_税抜.xlsx")
ws = wb["template1"]


# 見積data取得
## excel data >> 二次元リスト化

## Header Left
quote_header = []

for row in ws.iter_rows(min_col=3, max_col=78, max_row=29):
    for cell in row:
        cell_value = cell.value
        if cell_value is not None:
            quote_header.append(cell_value)

## Quote Header Index番号確認

#for i, j in enumerate(quote_header):
#    print('{0}:{1}'.format(i,j))



## 見積詳細

def is_empty(cell):
    return cell.value is None or not str(cell.value).strip()


quote_details = [] #見積書Detail list
quote_start = 32 #見積変動開始行

for i in ws.iter_rows(min_col=2, min_row=quote_start, max_col=78):
    if all(is_empty(c) for c in i):
        break
    values_original = [cell.value for cell in i]
    values_original_filtered = [j for j in values_original if j is not None]
    quote_details.append(values_original_filtered)

#for k, l in enumerate(quote_details):
#    print('{0}:{1}'.format(k,l))


### 見積項目数に応じてリスト分割

quote_numbers = len(quote_details)

dt_row_p1 = 29 #見積項目数上限Page.1
dt_row_p2 = 84 #見積項目数上限Page.2


dt_col_list = []
for k in quote_details:
    dt_col_list.append({
        "works":k[0],
        "volume" : '{:,.3f}'.format(k[2]),
        "unit_v" : k[3],
        "currency" : k[4],
        "unit_p" : '{:,.2f}'.format(k[5]),
        "amount" : '{:,.2f}'.format(k[6]),
        "ex_rate" : k[7],
        "amount_j" : '{:,}'.format(k[8]),
        "tax" : k[9],
        })
dt_list_1 = dt_col_list[0:dt_row_p1]
dt_list_2 = dt_col_list[dt_row_p1:dt_row_p2]
dt_list_3 = dt_col_list[dt_row_p2:quote_numbers]



tax_total_cal = []
duty_free_total_cal = []
for m in quote_details:
    if m[9] == '〇':
        tax_total_cal.append(m[8])
    else:
        duty_free_total_cal.append(m[8])

grand_total = []

duty_free_total = sum(duty_free_total_cal)
grand_total.append({"duty_free_total" : '{:,}'.format(duty_free_total)})

tax_total = sum(tax_total_cal)
grand_total.append({"tax_total" : '{:,}'.format(tax_total)})

gtotal = tax_total + duty_free_total
grand_total.append({"gtotal" : '{:,}'.format(gtotal)})

#print(grand_total)

grandtotal_1 = []
grandtotal_2 = []
grandtotal_3 = []

if quote_numbers <= dt_row_p1 - 4 or quote_numbers == dt_list_1:
    grandtotal_1 = grand_total
elif quote_numbers <= dt_row_p2 - 4:
    grandtotal_2 = grand_total
else:
    grandtotal_3 = grand_total


#print('-----------------')
#print(len(grandtotal_1))
#print('-----------------')
#print(len(grandtotal_2))
#print('-----------------')
#print(len(grandtotal_3))
#print('-----------------')


# Remarks

remarks = []
remarks_start = quote_start + quote_numbers + 6

for row in ws.iter_rows(min_col=4, min_row=remarks_start):
    for cell in row:
        if all(is_empty(c) for c in row):
            break
        cell_value = cell.value
        if cell_value is not None:
            remarks.append(cell_value)
#for k, l in enumerate(remarks):
#    print('{0}:{1}'.format(k,l))

remark1 = []
remark2 = []
remark3 = []


if len(grandtotal_1) > 0:
    remark_space1 = (124.35 - (len(dt_list_1) * 4.95 + 23.43))/3.62
    remark1 = remarks[0:int(remark_space1)]
    remark2 = remarks[int(remark_space1):len(remarks)]
elif len(grandtotal_2) > 0:
    remark2 = remarks
else:
    remark3 = remarks

# print('----------------')
# print(remark1)
# print('----------------')
# print(remark2)
# print('----------------')
# print(remark3)
# print('----------------')



## Excel 閉じる
wb.close()

